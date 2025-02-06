# С помощью питона создать excel файл.
# С помощью цикла и модуля random заполнить 2 колонки.
# Создать второй лист.
# На втором листе заполнить 3 колонки:
# 1 - дата
# 2 - порядковый номер в цикле
# 3 - uuid из модуля uuid

from openpyxl import Workbook
import random
import datetime as dt
import uuid

wb = Workbook()
ws = wb.active

# adding to excel file random numbers
columns_task_1 = ['A', 'B']
for column in columns_task_1:
    for row in range(1, 11):
        cell = column + str(row)
        ws[cell] = random.randint(1,100)

# 2nd sheet task
ws1 = wb.create_sheet("2nd sheet")
date = dt.datetime.now().strftime("%d-%m-%Y")
for i in range(1, 21):
    column_A = 'A' + str(i)
    column_B = 'B' + str(i)
    column_C = 'C' + str(i)
    ws1[column_A] = date
    ws1[column_B] = i
    ws1[column_C] = str(uuid.uuid4())

wb.save('example.xlsx')
